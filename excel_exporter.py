from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# -----------------------------
# Common Sheet Styling
# -----------------------------
def style_sheet(sheet):

    header_fill = PatternFill(
        fill_type="solid",
        start_color="4F81BD",
        end_color="4F81BD"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    for cell in sheet[1]:

        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

    sheet.freeze_panes = "A2"

    sheet.auto_filter.ref = sheet.dimensions


def auto_fit(sheet):

    for column in sheet.columns:

        length = 0

        column_letter = get_column_letter(column[0].column)

        for cell in column:

            if cell.value:

                length = max(
                    length,
                    len(str(cell.value))
                )

            cell.alignment = Alignment(
                wrap_text=True,
                vertical="top"
            )

        sheet.column_dimensions[column_letter].width = min(length + 4, 50)

def create_summary_sheet(workbook, results):

    sheet = workbook.active

    sheet.title = "Summary"

    sheet.append([
        "Requirement",
        "Status"
    ])

    for result in results:

        sheet.append([
            result["requirement"],
            result["status"]
        ])

    style_sheet(sheet)

    auto_fit(sheet)

def export_to_excel(results, file_path):

    workbook = Workbook()

    create_summary_sheet(workbook, results)

    create_ba_questions_sheet(workbook, results)

    create_test_scenarios_sheet(workbook, results)

    create_positive_test_cases_sheet(workbook, results)

    create_negative_test_cases_sheet(workbook, results)

    create_boundary_test_cases_sheet(workbook, results)

    create_risks_sheet(workbook, results)

    workbook.save(file_path)

def create_ba_questions_sheet(workbook, results):

    sheet = workbook.create_sheet("BA Questions")

    sheet.append([
        "Requirement",
        "BA Question ID",
        "Question"
    ])

    for result in results:

        if result["status"] == "Failed":
            continue

        data = result["data"]

        for question in data["ba_questions"]:

            sheet.append([
                result["requirement"],
                question["id"],
                question["question"]
            ])

    style_sheet(sheet)

    auto_fit(sheet)

def create_test_scenarios_sheet(workbook, results):

    sheet = workbook.create_sheet("Test Scenarios")

    sheet.append([
        "Requirement",
        "Scenario ID",
        "Description"
    ])

    for result in results:

        if result["status"] == "Failed":
            continue

        data = result["data"]

        for scenario in data["test_scenarios"]:

            sheet.append([
                result["requirement"],
                scenario["id"],
                scenario["description"]
            ])

    style_sheet(sheet)

    auto_fit(sheet)

def create_positive_test_cases_sheet(workbook, results):

    sheet = workbook.create_sheet("Positive Test Cases")

    sheet.append([
        "Requirement",
        "Test Case ID",
        "Description",
        "Steps",
        "Expected Result"
    ])

    for result in results:

        if result["status"] == "Failed":
            continue

        data = result["data"]

        for tc in data["positive_test_cases"]:

            steps = "\n".join(
                f"{i+1}. {step}"
                for i, step in enumerate(tc["steps"])
            )

            sheet.append([
                result["requirement"],
                tc["id"],
                tc["description"],
                steps,
                tc["expected_result"]
            ])

    style_sheet(sheet)

    auto_fit(sheet)

def create_negative_test_cases_sheet(workbook, results):

    sheet = workbook.create_sheet("Negative Test Cases")

    sheet.append([
        "Requirement",
        "Test Case ID",
        "Description",
        "Steps",
        "Expected Result"
    ])

    for result in results:

        if result["status"] == "Failed":
            continue

        data = result["data"]

        for tc in data["negative_test_cases"]:

            steps = "\n".join(
                f"{i+1}. {step}"
                for i, step in enumerate(tc["steps"])
            )

            sheet.append([
                result["requirement"],
                tc["id"],
                tc["description"],
                steps,
                tc["expected_result"]
            ])

    style_sheet(sheet)

    auto_fit(sheet)

def create_boundary_test_cases_sheet(workbook, results):

    sheet = workbook.create_sheet("Boundary Test Cases")

    sheet.append([
        "Requirement",
        "Test Case ID",
        "Description",
        "Steps",
        "Expected Result"
    ])

    for result in results:

        if result["status"] == "Failed":
            continue

        data = result["data"]

        for tc in data["boundary_test_cases"]:

            steps = "\n".join(
                f"{i+1}. {step}"
                for i, step in enumerate(tc["steps"])
            )

            sheet.append([
                result["requirement"],
                tc["id"],
                tc["description"],
                steps,
                tc["expected_result"]
            ])

    style_sheet(sheet)

    auto_fit(sheet)

def create_risks_sheet(workbook, results):

    sheet = workbook.create_sheet("Risks")

    sheet.append([
        "Requirement",
        "Risk ID",
        "Description"
    ])

    for result in results:

        if result["status"] == "Failed":
            continue

        data = result["data"]

        for risk in data["risks"]:

            sheet.append([
                result["requirement"],
                risk["id"],
                risk["description"]
            ])

    style_sheet(sheet)

    auto_fit(sheet)  