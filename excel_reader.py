from openpyxl import load_workbook

def read_requirements(file):

    workbook = load_workbook(file)
    sheet = workbook.active

    requirements = []
    skipped_rows = 0

    # Skip header row
    for row in sheet.iter_rows(min_row=2, values_only=True):

        requirement = row[0]

        if requirement is None or str(requirement).strip() == "":
            skipped_rows += 1
            continue

        requirements.append(str(requirement).strip())

    return requirements, skipped_rows